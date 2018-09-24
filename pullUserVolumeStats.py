from openpyxl import load_workbook
import xlsxwriter
import datetime, numpy as np
import calendar
import os
import time

dayInSeconds = 60*60*24
userVolumeStatsFileName = "statistic_id253577_instagram_-number-of-monthly-active-users-2013-2018.xlsx"
userVolumeStatsRelevantSheet = "Data"

def convert_str_to_date_notime(datetimeInStringFormat):
    strr = datetimeInStringFormat.split(" ")[0]
    dateArray = strr.split("-")
    year = dateArray[0]
    month = dateArray[1]
    day = dateArray[2]
    monthStrAbbrv = calendar.month_abbr[int(month)]
    if day == "0" or day == "00":
        day = "01"
    item3 = datetime.datetime.strptime('{0} {1} {2}'.format(monthStrAbbrv, day, year), '%b %d %Y')
    item3 = item3.date()
    return item3


def customConvertStringToDate(monthStrAbbrv, yearStr, dayStr=False):
    if not dayStr:
        dayStr = "01"
    strAsDatetime = datetime.datetime.strptime('{0} {1} {2}'.format(monthStrAbbrv, dayStr, yearStr), '%b %d %Y')
    date = strAsDatetime.date()
    return date

def toTimestamp(d):
    if type(d) == type(str("a")):
        date_ = customConvertStringToDate(d[:3], "20"+d[-2:])
    else:
        date_ = d
    return calendar.timegm(date_.timetuple())

def timestampToStr(timestampInt):
    if type(timestampInt) == type(str("string")):
        try:
            timestampInt = int(timestampInt)
        except ValueError:
            pass
    if type(timestampInt) == type(str("string")):
        try:
            timestampInt = convert_str_to_date_notime(timestampInt)
            timestampInt = toTimestamp(timestampInt)
        except TypeError:
            pass


    return datetime.datetime.utcfromtimestamp(timestampInt).strftime('%Y-%m-%d %H:%M:%S')

def checkIfModifiedSheetExists():
    return os.path.isfile(os.getcwd() + "/ModifiedUserVolume.xlsx")

def checkForUnmotifiedSheet():
    return os.path.isfile(os.getcwd()+"/statistic_id253577_instagram_-number-of-monthly-active-users-2013-2018.xlsx")

def createModifiedUserVolumeSpreadsheet(fullDatesArray, fullValuesArray):
    # First we'll create the spreadsheet if it does not already exist.
    if not checkIfModifiedSheetExists():
        workbook = xlsxwriter.Workbook('ModifiedUserVolume.xlsx', {'default_date_format':'dd/mm/yy'})
        worksheet = workbook.add_worksheet("ModifiedData")
    # If it does exist, we'll delete it and crate a new blank one.
    else:
        os.remove('ModifiedUserVolume.xlsx')
        workbook = xlsxwriter.Workbook('ModifiedUserVolume.xlsx')
        worksheet = workbook.add_worksheet("ModifiedData")
    # Now we'll write everything
    for i in range(len(fullDatesArray)):
        # Python starts counting at 0, but excel starts counting at 1
        currentRowString = str(i+1)
        # Excel gives us a weird format when we try to write the datetime directly, so we convert it to unixTimestamp
        unixTimestamp = int(time.mktime(fullDatesArray[i].timetuple()))
        # Now it's just write away. Dates (as unix) first
        worksheet.write('A'+currentRowString, unixTimestamp)
        # Now or the IG user volume, converted from millions to plain numbers
        worksheet.write('B'+currentRowString, round(fullValuesArray[i]*1000000), 0)
    workbook.close()
    return



def returnDailyInterpolatedArray(startDT, endDT):
    dayInterpolatedListStr = list()
    dayInterpolatedListDates = list()
    if type(startDT) == type(str("a")) and type(endDT)== type(str("a")):
        startDT1 = customConvertStringToDate(startDT[:3], "20"+ startDT[-2:])
        endDT1 = customConvertStringToDate(endDT[:3], "20"+ endDT[-2:])
        startTS, endTS = toTimestamp(startDT1), toTimestamp(endDT1)
    else:
        startTS, endTS = toTimestamp(startDT), toTimestamp(endDT)
    dayInterpolatedArrayTS = np.arange(startTS, endTS, dayInSeconds)
    for item in dayInterpolatedArrayTS:
        dayInterpolatedListStr.append(timestampToStr(item))
    for otheritem in dayInterpolatedListStr:
        dayInterpolatedListDates.append(convert_str_to_date_notime(otheritem))
    return dayInterpolatedListDates


def interpBetweenFirstLastDays(firstDateCell, lastDateCell):
    allDates = returnDailyInterpolatedArray(firstDateCell, lastDateCell)
    allDates.append(customConvertStringToDate(lastDateCell[:3], "20"+lastDateCell[-2:]))
    return allDates




def returnInterpolatedValuesByDayArray(startVal, endVal, startdateCell, enddateCell):
    separatedDailyDates = returnDailyInterpolatedArray(startdateCell, enddateCell)
    return np.arange(startVal, endVal, (endVal-startVal) / len(separatedDailyDates))



def createFullDatesArray(statistaExcelFileName, statistaExcelSheetName):
    workbook = load_workbook(statistaExcelFileName, data_only=True)
    sheet = workbook[statistaExcelSheetName]
    return interpBetweenFirstLastDays(sheet["B" + str(6)].value, sheet["B" + str(17)].value)

def createFullValuesArray(statistaExcelFileName, statistaExcelSheetName):
    workbook = load_workbook(statistaExcelFileName, data_only=True)
    sheet = workbook[statistaExcelSheetName]
    allVals = list()
    allDates = list()
    # First we interpolate all the data provided by Statista
    for i in range(6, 17):
        startDateCellStr = "B" + str(i)
        strt = sheet[startDateCellStr].value
        endDateCellStr = "B" + str(i+1)
        end = sheet[endDateCellStr].value
        startVal1 = sheet["C" + str(i)].value
        endVal1 = sheet["C" + str(i+1)].value
        if i == 16:
            daysArry = interpBetweenFirstLastDays(strt, end)
            allDates += daysArry
            almostAllValsThisIter = returnInterpolatedValuesByDayArray(startVal1, endVal1, strt, end)
            allValsThisIter = np.append(almostAllValsThisIter, endVal1)
            allVals += allValsThisIter.tolist()
        else:
            daysArry = interpBetweenFirstLastDays(strt, end)[:-1]
            allDates += daysArry
            allValsThisIter = returnInterpolatedValuesByDayArray(startVal1, endVal1, strt, end)
            allVals += allValsThisIter.tolist()
    # After an error I realized that I was trying to scale posts before the inital Statista given date of 1/1/13
    # There were definitely users before that since Instagram launched October 6, 2010.
    # So I'm just going to interpolate linearly from launch date to January 1st, 2013. Let's create it
    launchDate_DT = customConvertStringToDate("Oct", "2010", "6")
    initialStatista_DT = customConvertStringToDate("Jan", "2013", "1")
    launchVals = 0
    statistaValsInMil = 90
    beforeStatistaDates = returnDailyInterpolatedArray(launchDate_DT, initialStatista_DT)
    beforeStatistaUsers = returnInterpolatedValuesByDayArray(launchVals, statistaValsInMil, launchDate_DT, initialStatista_DT)
    # So Second, let's add all that data from Oct 6, 2010 to Jan 1, 13.
    finalAllDate = beforeStatistaDates[:] + allDates[:]
    finalAllVals_inMil = beforeStatistaUsers[:].tolist() + allVals[:]
    # I also need to set all the values up to today...
    return finalAllDate, finalAllVals_inMil

dts, _vals = createFullValuesArray(userVolumeStatsFileName, userVolumeStatsRelevantSheet)
createModifiedUserVolumeSpreadsheet(dts, _vals)
dtsNP = [int(time.mktime(elem.timetuple())) for elem in dts]

timeStampValueDict = dict(zip(dtsNP, _vals))
dateValueDict = dict(zip(dts, _vals))

# count = 0
#
# for k in timeStampValueDict:
#     count += 1
#     print(k, timeStampValueDict[k])
#     if count > 100:
#         break

